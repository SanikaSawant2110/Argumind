import os
import time
from fastapi.responses import StreamingResponse
import io

AUTO_EXPORT_PATH = os.path.join(os.getcwd(), "argumind_debates.xlsx")


def _build_workbook(history: list):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    green_font  = Font(bold=True, color="22C55E", size=10)
    red_font    = Font(bold=True, color="EF4444", size=10)
    thin_border = Border(bottom=Side(style="thin", color="333355"))

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Debate Summary"

    model_cols = sorted({
        k for row in history
        for k in row
        if k.endswith("_score") and k != "hallucination_score"
    })

    summary_headers = [
        "#", "Timestamp", "Topic / Query", "Winner", "Confidence %",
        "PRO Votes", "CON Votes", "Most Hallucinating Model",
    ] + [c.replace("_score", "").replace("_", " ").title() + " Score" for c in model_cols]

    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    col_widths = [5, 18, 50, 10, 14, 10, 10, 32] + [18] * len(model_cols)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r_idx, row in enumerate(history, 2):
        ts = row.get("timestamp", "")
        if isinstance(ts, (int, float)) and ts > 1e9:
            ts = time.strftime("%Y-%m-%d %H:%M", time.localtime(ts))

        conf = row.get("confidence", 0)
        conf_pct = round(conf * 100 if conf <= 1 else conf, 1)

        values = [
            r_idx - 1, ts, row.get("query", ""),
            row.get("final_decision", ""), conf_pct,
            row.get("pro_votes", ""), row.get("con_votes", ""),
            row.get("most_hallucinating_model", ""),
        ] + [round(row.get(c, 0), 4) for c in model_cols]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")
            cell.border = thin_border
            cell.font = Font(size=10)
            if col == 4:
                cell.font = green_font if val == "PRO" else red_font
        ws.row_dimensions[r_idx].height = 18

    # ── Sheet 2: Arguments ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Arguments & Rebuttals")
    arg_headers = [
        "#", "Topic",
        "PRO 1", "PRO 2", "PRO 3", "PRO 4", "PRO 5",
        "CON 1", "CON 2", "CON 3", "CON 4", "CON 5",
        "PRO Rebuttal 1", "PRO Rebuttal 2", "PRO Rebuttal 3", "PRO Rebuttal 4",
        "CON Rebuttal 1", "CON Rebuttal 2", "CON Rebuttal 3", "CON Rebuttal 4",
    ]
    for col, h in enumerate(arg_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="0f3460")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 40
    for i in range(3, len(arg_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 35

    for r_idx, row in enumerate(history, 2):
        pro_args = row.get("pro_arguments", []) or []
        con_args = row.get("con_arguments", []) or []
        pro_reb  = (row.get("rebuttals") or {}).get("pro_rebuttals", []) or []
        con_reb  = (row.get("rebuttals") or {}).get("con_rebuttals", []) or []
        def _g(lst, i): return lst[i] if i < len(lst) else ""
        values = [r_idx - 1, row.get("query", "")] + \
                 [_g(pro_args, i) for i in range(5)] + \
                 [_g(con_args, i) for i in range(5)] + \
                 [_g(pro_reb, i) for i in range(4)] + \
                 [_g(con_reb, i) for i in range(4)]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=r_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=9)
            cell.border = thin_border
        ws2.row_dimensions[r_idx].height = 60

    # ── Sheet 3: Judge Verdicts ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Judge Verdicts")
    judge_headers = [
        "#", "Topic",
        "Gemini Verdict", "Gemini Conf %",
        "OpenRouter Verdict", "OpenRouter Conf %",
        "Heuristic Verdict", "Heuristic Conf %",
        "Consensus Winner", "Consensus Conf %",
        "PRO Votes", "CON Votes", "Reasoning",
    ]
    for col, h in enumerate(judge_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="16213e")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for w, col in zip([5, 40, 14, 14, 18, 18, 18, 18, 16, 16, 10, 10, 60],
                      range(1, len(judge_headers) + 1)):
        ws3.column_dimensions[get_column_letter(col)].width = w

    for r_idx, row in enumerate(history, 2):
        ivs = row.get("individual_verdicts", [])
        def _iv(i, key, default=""): return ivs[i].get(key, default) if i < len(ivs) else default
        def conf_pct(c):
            try:
                c = float(c)
                return round(c * 100 if c <= 1 else c, 1)
            except Exception:
                return 65.0
        values = [
            r_idx - 1, row.get("query", ""),
            _iv(0, "winner"), conf_pct(_iv(0, "confidence", 0.65)),
            _iv(1, "winner"), conf_pct(_iv(1, "confidence", 0.65)),
            _iv(2, "winner"), conf_pct(_iv(2, "confidence", 0.65)),
            row.get("final_decision", ""), conf_pct(row.get("confidence", 0.65)),
            row.get("pro_votes", ""), row.get("con_votes", ""),
            row.get("reasoning", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=r_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col not in (2, 13) else "left", wrap_text=True, vertical="top")
            cell.font = Font(size=9)
            cell.border = thin_border
            if col in (3, 5, 7, 9):
                cell.font = green_font if val == "PRO" else (red_font if val == "CON" else Font(size=9))
        ws3.row_dimensions[r_idx].height = 45

    # ── Sheet 4: Hallucination Scores ─────────────────────────────────────────
    ws4 = wb.create_sheet("Hallucination Scores")
    hall_headers = ["#", "Topic", "Model", "Hallucination Score", "Consensus Similarity", "Risk Level"]
    for col, h in enumerate(hall_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for w, col in zip([5, 40, 30, 20, 22, 15], range(1, 7)):
        ws4.column_dimensions[get_column_letter(col)].width = w

    hall_row = 2
    for r_idx, row in enumerate(history, 1):
        for score in (row.get("hallucination_scores") or []):
            hs = score.get("hallucination_score", 0)
            risk = "High" if hs > 0.25 else "Medium" if hs > 0.12 else "Low"
            values = [r_idx, row.get("query", ""), score.get("model", ""), hs, score.get("consensus_similarity", 0), risk]
            for col, val in enumerate(values, 1):
                cell = ws4.cell(row=hall_row, column=col, value=val)
                cell.alignment = Alignment(horizontal="center" if col != 3 else "left")
                cell.font = Font(size=9)
                cell.border = thin_border
                if col == 6:
                    cell.font = Font(
                        bold=True, size=9,
                        color="EF4444" if risk == "High" else "F59E0B" if risk == "Medium" else "22C55E"
                    )
            hall_row += 1

    return wb


def auto_save_excel(history: list):
    try:
        wb = _build_workbook(history)
        wb.save(AUTO_EXPORT_PATH)
        print(f"  💾 Excel auto-saved → {AUTO_EXPORT_PATH}")
    except Exception as e:
        print(f"  ⚠️ Excel auto-save failed: {e}")


def build_excel(history: list):
    if not history:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No Data"
        ws["A1"] = "No debates run yet."
    else:
        wb = _build_workbook(history)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=argumind_debates.xlsx"},
    )